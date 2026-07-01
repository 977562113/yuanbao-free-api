from openai import OpenAI
import json
import re
import time
import random
import openpyxl

# python test-stand-table-output-readfile.py
# 功能：从 data/20260630.xlsx 读取股票列表，每次查询10条，
#       随机间隔15-30秒，将"央国企"和"社保持仓"结果回写到 Excel

# ============ 配置 ============
API_URL = "http://localhost:8000/v1"
API_KEY = "sk-your-api-key-here"  # 使用 .env 中配置的 API Key
EXCEL_PATH = "data/20260630.xlsx"
BATCH_SIZE = 10          # 每次查询股票数
SLEEP_MIN = 15           # 最小间隔秒数
SLEEP_MAX = 30           # 最大间隔秒数

# 新增列的列名
COL_YANGGUOQI = "央国企"
COL_SHEBAO = "社保持仓"

# ============ 创建 OpenAI 客户端 ============
client = OpenAI(base_url=API_URL, api_key=API_KEY)


def load_stocks(excel_path):
    """从 Excel 读取股票列表，返回 [(row_index, 代码, 名称), ...]"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    stocks = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, 1).value   # 代码
        name = ws.cell(row, 2).value   # 名称
        if name:
            stocks.append((row, str(code).strip(), str(name).strip()))
    return wb, ws, stocks


def ensure_columns(ws):
    """确保 Excel 中存在央国企和社保持仓列，返回对应列号"""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if COL_YANGGUOQI in headers:
        col_ygq = headers.index(COL_YANGGUOQI) + 1
    else:
        col_ygq = ws.max_column + 1
        ws.cell(1, col_ygq, COL_YANGGUOQI)
    if COL_SHEBAO in headers:
        col_sb = headers.index(COL_SHEBAO) + 1
    else:
        col_sb = col_ygq + 1
        ws.cell(1, col_sb, COL_SHEBAO)
    return col_ygq, col_sb


def query_stocks(stock_names):
    """向 API 发送一批股票名称，返回解析后的列表"""
    name_str = ", ".join(stock_names)
    prompt = (
        "1.要求：根据用户提供股票列表，分析是否央国企，是否社保持仓股"
        f"2.股票列表：{name_str}"
        "3.输出：以 JSON 数组格式返回，每个元素包含名称、央国企、社保持仓字段"
        "4.示例：[{\"名称\":\"浦发银行\",\"央国企\":\"是\",\"社保持仓\":\"否\"}]"
        "5.只返回 JSON 数组，严格按照示例格式返回JSON，不要其他内容(否则干扰 JSON 解析)"
    )

    stream = client.chat.completions.create(
        model="deepseek-v3-search",
        messages=[{"role": "", "content": prompt}],
        stream=True,
    )

    full_text = ""
    for chunk in stream:
        if not chunk.choices:
            continue
        for choice_index, choice in enumerate(chunk.choices):
            delta = choice.delta
            content = delta.content if delta.content else ""
            finish_reason = choice.finish_reason
            if finish_reason:
                print(f"  流式响应结束，原因: {finish_reason}")
            if content:
                try:
                    json_obj = json.loads(content)
                    msg_type = json_obj.get("type", "")
                    if msg_type == "text":
                        msg = json_obj.get("msg", "")
                        if msg and not msg.startswith("[](@replace"):
                            full_text += msg
                    elif msg_type == "mark":
                        mark_obj = json_obj.get("mark", {})
                        if isinstance(mark_obj, dict):
                            msg = mark_obj.get("content", "")
                            if msg:
                                full_text += msg
                    else:
                        for key in ["content", "text", "data", "message"]:
                            if key in json_obj:
                                value = json_obj[key]
                                if isinstance(value, str):
                                    full_text += value
                                break
                except json.JSONDecodeError:
                    full_text += content
                except Exception as e:
                    print(f"  其他错误: {type(e).__name__}: {e}")
                    full_text += content

    # 清理引用标注
    full_text = re.sub(r'\[\]\(@mark_underline=(\d+)\)', r'[\1]', full_text)
    full_text = re.sub(r'\[citation:\d+\]', '', full_text)
    full_text = re.sub(r'\[(\d+)\]', '', full_text)

    # 从 full_text 中提取 JSON 数组
    results = parse_json_from_text(full_text)
    return results


def parse_json_from_text(text):
    """从返回文本中提取 JSON 数组"""
    # 尝试直接解析
    try:
        data = json.loads(text.strip())
        if isinstance(data, list):
            return data
    except Exception:
        pass
    # 尝试用正则提取 JSON 数组
    match = re.search(r'\[[\s\S]*\]', text)
    if match:
        try:
            return json.loads(match.group())
        except Exception:
            pass
    print(f"  ⚠️ 无法解析 JSON，原始文本前200字：{text[:200]}")
    return []


def main():
    print(f"正在读取 Excel 文件：{EXCEL_PATH}")
    wb, ws, stocks = load_stocks(EXCEL_PATH)
    col_ygq, col_sb = ensure_columns(ws)
    print(f"共读取 {len(stocks)} 条股票，央国企列={col_ygq}，社保持仓列={col_sb}")

    # 分批处理
    total_batches = (len(stocks) + BATCH_SIZE - 1) // BATCH_SIZE
    for batch_idx in range(total_batches):
        start = batch_idx * BATCH_SIZE
        end = min(start + BATCH_SIZE, len(stocks))
        batch = stocks[start:end]
        batch_names = [name for _, _, name in batch]

        print(f"\n[{batch_idx+1}/{total_batches}] 查询：{', '.join(batch_names)}")

        # 查询前随机睡眠（第一批除外）
        if batch_idx > 0:
            sleep_sec = random.randint(SLEEP_MIN, SLEEP_MAX)
            print(f"  等待 {sleep_sec} 秒...")
            time.sleep(sleep_sec)

        try:
            results = query_stocks(batch_names)
        except Exception as e:
            print(f"  ❌ 查询失败：{e}")
            continue

        # 构建 名称 -> 结果 的映射
        result_map = {}
        for item in results:
            name = item.get("名称", "").strip()
            if name:
                result_map[name] = item

        # 回写到 Excel 对应行
        updated = 0
        for row, code, name in batch:
            item = result_map.get(name)
            if item:
                ws.cell(row, col_ygq, item.get("央国企", ""))
                ws.cell(row, col_sb, item.get("社保持仓", ""))
                updated += 1
                print(f"  ✓ {name}：央国企={item.get('央国企','')}, 社保持仓={item.get('社保持仓','')}")
            else:
                print(f"  ✗ {name}：未找到匹配结果")

        print(f"  本批更新 {updated}/{len(batch)} 条")

        # 每批处理后保存一次（防止中断丢失）
        wb.save(EXCEL_PATH)
        print(f"  已保存到 {EXCEL_PATH}")

    print(f"\n✅ 全部完成！结果已保存到 {EXCEL_PATH}")


if __name__ == "__main__":
    main()
