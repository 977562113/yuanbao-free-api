from openai import OpenAI
import json
import re
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import os
import time
import random

# 配置
API_URL = "http://localhost:8000/v1"
API_KEY = "sk-your-api-key-here"  # 使用 .env 中配置的 API Key
INPUT_FILE = "data/筛选结果_120天跌15%以上_分位25%以下_非ST_非688_20260630.xlsx"
OUTPUT_FILE = "data/result.xlsx"
BATCH_SIZE = 10

# 创建 OpenAI 客户端
client = OpenAI(base_url=API_URL, api_key=API_KEY)


def read_stock_info(file_path):
    """从 Excel 读取个股代码和名称列表，返回 [(code, name), ...]"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    stocks = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            code = str(row[0]).strip()
            name = str(row[1]).strip()
            stocks.append((code, name))
    return stocks


def parse_response(text):
    """解析响应文本，支持 JSON 对象格式（键为股票代码）"""
    text = text.strip()
    
    # 尝试提取 JSON 对象
    # 查找 { 和 } 之间的内容
    start = text.find('{')
    end = text.rfind('}')
    
    if start != -1 and end != -1 and end > start:
        json_str = text[start:end + 1]
        try:
            data = json.loads(json_str)
            if isinstance(data, dict):
                # 转换为列表格式，每个元素包含代码字段
                result = []
                for code, info in data.items():
                    if isinstance(info, dict):
                        info['代码'] = code
                        result.append(info)
                return result
        except json.JSONDecodeError:
            pass
    
    # 尝试 JSON 数组格式
    start = text.find('[')
    end = text.rfind(']')
    
    if start != -1 and end != -1 and end > start:
        json_str = text[start:end + 1]
        try:
            data = json.loads(json_str)
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass
    
    # 如果不是 JSON，尝试作为 markdown 表格解析
    return parse_markdown_table(text)


def parse_markdown_table(text):
    """解析 markdown 表格文本，返回行列表（每行为字典列表）"""
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    # 找到表头行（包含 | 的行）
    table_lines = [l for l in lines if '|' in l]
    if len(table_lines) < 3:
        return []

    # 第一行是表头，第二行是分隔符，后面是数据
    header_line = table_lines[0]
    headers = [h.strip() for h in header_line.split('|') if h.strip()]

    rows = []
    for line in table_lines[2:]:  # 跳过分隔符行
        cells = [c.strip() for c in line.split('|') if c.strip()]
        if len(cells) >= len(headers):
            cells = cells[:len(headers)]
        elif len(cells) < len(headers):
            cells += [''] * (len(headers) - len(cells))
        rows.append(dict(zip(headers, cells)))
    return rows


def query_stocks(stock_info):
    """发送 API 请求查询股票信息，stock_info 为 [(code, name), ...]"""
    # 构建股票列表字符串：代码(名称)
    stock_list = [f"{code}({name})" for code, name in stock_info]
    stock_str = "，".join(stock_list)
    prompt = (
        "1.要求：根据用户提供股票列表，分析是否央国企，是否社保持仓股，合理估值价格，是否推荐购买，120日涨跌幅，120日价格分位，所属行业，现价，120日前价格，120日价格涨幅，120日价格涨幅分位；"
        f"2.{stock_str}；"
        "3.输出：以 JSON 对象格式返回，键为股票代码，值为包含分析结果的字典（包含名称字段）；"
        "4.示例：{\"600000\":{\"名称\":\"浦发银行\",\"央国企\":\"是\",\"社保持仓\":\"否\",\"合理估值价格\":\"10元\",\"是否推荐购买\":\"增持\",\"120日涨跌幅\":\"-15%\",\"120日价格分位\":\"20%\",\"所属行业\":\"银行\",\"现价\":\"8.5元\",\"120日前价格\":\"10元\",\"120日价格涨幅\":\"-15%\",\"120日价格涨幅分位\":\"25%\"}}；"
        "5.只返回 JSON 对象，不要其他内容"
    )

    print(f"\n正在查询: {stock_str}")
    stream = client.chat.completions.create(
        model="deepseek-v3-search",
        messages=[{"role": "", "content": prompt}],
        stream=True
    )

    full_text = ""
    for chunk in stream:
        if not chunk.choices:
            continue
        for choice in chunk.choices:
            delta = choice.delta
            content = delta.content if delta.content else ""
            if choice.finish_reason:
                print(f"  流式响应结束，原因: {choice.finish_reason}")
            if content:
                try:
                    json_obj = json.loads(content)
                    msg_type = json_obj.get("type", "")
                    if msg_type == "text":
                        msg = json_obj.get("msg", "")
                        if msg and not msg.startswith("[](@replace"):
                            full_text += msg
                    elif msg_type == "mark":
                        mark_obj = json_obj.get("mark", {})
                        if isinstance(mark_obj, dict):
                            msg = mark_obj.get("content", "")
                            if msg:
                                full_text += msg
                    else:
                        print(f"  ⚠️ 未处理的消息类型: {msg_type}")
                except json.JSONDecodeError:
                    full_text += content
                except Exception as e:
                    print(f"  其他错误: {type(e).__name__}: {e}")
                    full_text += content

    # 清理引用标签
    full_text = re.sub(r'\[\]\(@mark_underline=(\d+)\)', r'[\1]', full_text)
    full_text = re.sub(r'\[citation:\d+\]', '', full_text)
    full_text = re.sub(r'\[(\d+)\]', '', full_text)

    print(f"  响应字符数: {len(full_text)}")
    return full_text


def update_source_file(result_rows, source_file):
    """将查询结果更新到源文件，新增"央国企"、"社保持仓"、"合理估值价格"、"是否推荐购买"、"120日涨跌幅"、"120日价格分位"、"是否已发公告"列"""
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    
    # 读取表头
    headers = [c.value for c in ws[1]]
    
    # 找到代码列索引
    code_col_idx = None
    for idx, h in enumerate(headers):
        if h and '代' in str(h) and '码' in str(h):
            code_col_idx = idx
            break
    if code_col_idx is None:
        print("❌ 未找到代码列")
        return
    
    # 构建代码到行号的映射
    code_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=code_col_idx + 1).value
        if code:
            code_to_row[str(code).strip()] = row_idx
    
    # 检查/新增列
    new_cols = ["央国企", "社保持仓", "合理估值价格", "是否推荐购买", "120日涨跌幅", "120日价格分位", "所属行业", "现价", "120日前价格", "120日价格涨幅", "120日价格涨幅分位"]
    col_indices = {}  # 列名 -> 列索引(0-based)
    
    for col_name in new_cols:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name)
        else:
            # 在末尾新增列
            new_idx = len(headers)
            headers.append(col_name)
            ws.cell(row=1, column=new_idx + 1, value=col_name)
            col_indices[col_name] = new_idx
            print(f"  新增列: {col_name} (第{new_idx + 1}列)")
    
    # 处理带下拉选项的列
    dropdown_cols = [
        ("是否已发公告", "是,否,待定", "待定"),
        ("是否买入", "是,否,待定", "待定"),
    ]
    
    for col_name, options, default in dropdown_cols:
        if col_name in headers:
            col_idx = headers.index(col_name)
        else:
            col_idx = len(headers)
            headers.append(col_name)
            ws.cell(row=1, column=col_idx + 1, value=col_name)
            print(f"  新增列: {col_name} (第{col_idx + 1}列)")
        col_indices[col_name] = col_idx
        
        # 添加下拉选项（数据验证）
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        dv = DataValidation(
            type="list",
            formula1=f'"{options}"',
            allow_blank=True,
            showDropDown=False
        )
        dv.error = "请选择：" + options.replace(",", "、")
        dv.errorTitle = "输入无效"
        dv.prompt = f"请选择{col_name}"
        dv.promptTitle = col_name
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
        
        # 设置列宽
        ws.column_dimensions[col_letter].width = 18
    
    # 处理"备注"列（普通文本列）
    remark_col = "备注"
    if remark_col in headers:
        remark_col_idx = headers.index(remark_col)
    else:
        remark_col_idx = len(headers)
        headers.append(remark_col)
        ws.cell(row=1, column=remark_col_idx + 1, value=remark_col)
        print(f"  新增列: {remark_col} (第{remark_col_idx + 1}列)")
    col_indices[remark_col] = remark_col_idx
    
    # 设置备注列宽
    remark_col_letter = openpyxl.utils.get_column_letter(remark_col_idx + 1)
    ws.column_dimensions[remark_col_letter].width = 30
    
    # 处理"当日买入价格"列（数字列）
    buy_price_col = "当日买入价格"
    if buy_price_col in headers:
        buy_price_col_idx = headers.index(buy_price_col)
    else:
        buy_price_col_idx = len(headers)
        headers.append(buy_price_col)
        ws.cell(row=1, column=buy_price_col_idx + 1, value=buy_price_col)
        print(f"  新增列: {buy_price_col} (第{buy_price_col_idx + 1}列)")
    col_indices[buy_price_col] = buy_price_col_idx
    
    # 设置当日买入价格列宽
    buy_price_col_letter = openpyxl.utils.get_column_letter(buy_price_col_idx + 1)
    ws.column_dimensions[buy_price_col_letter].width = 15
    
    # 处理"买入后涨跌 (%)"列（百分比列）
    profit_col = "买入后涨跌(%)"
    if profit_col in headers:
        profit_col_idx = headers.index(profit_col)
    else:
        profit_col_idx = len(headers)
        headers.append(profit_col)
        ws.cell(row=1, column=profit_col_idx + 1, value=profit_col)
        print(f"  新增列: {profit_col} (第{profit_col_idx + 1}列)")
    col_indices[profit_col] = profit_col_idx
    
    # 设置买入后涨跌列宽
    profit_col_letter = openpyxl.utils.get_column_letter(profit_col_idx + 1)
    ws.column_dimensions[profit_col_letter].width = 15
    
    # 处理"买入时间"列（日期列）
    buy_time_col = "买入时间"
    if buy_time_col in headers:
        buy_time_col_idx = headers.index(buy_time_col)
    else:
        buy_time_col_idx = len(headers)
        headers.append(buy_time_col)
        ws.cell(row=1, column=buy_time_col_idx + 1, value=buy_time_col)
        print(f"  新增列: {buy_time_col} (第{buy_time_col_idx + 1}列)")
    col_indices[buy_time_col] = buy_time_col_idx
    
    # 设置买入时间列宽
    buy_time_col_letter = openpyxl.utils.get_column_letter(buy_time_col_idx + 1)
    ws.column_dimensions[buy_time_col_letter].width = 15
    
    # 处理"买入后最高价(%)"列（百分比列）
    max_price_col = "买入后最高价(%)"
    if max_price_col in headers:
        max_price_col_idx = headers.index(max_price_col)
    else:
        max_price_col_idx = len(headers)
        headers.append(max_price_col)
        ws.cell(row=1, column=max_price_col_idx + 1, value=max_price_col)
        print(f"  新增列: {max_price_col} (第{max_price_col_idx + 1}列)")
    col_indices[max_price_col] = max_price_col_idx
    
    # 设置买入后最高价列宽
    max_price_col_letter = openpyxl.utils.get_column_letter(max_price_col_idx + 1)
    ws.column_dimensions[max_price_col_letter].width = 18
    
    # 处理"买入后最低价(%)"列（百分比列）
    min_price_col = "买入后最低价(%)"
    if min_price_col in headers:
        min_price_col_idx = headers.index(min_price_col)
    else:
        min_price_col_idx = len(headers)
        headers.append(min_price_col)
        ws.cell(row=1, column=min_price_col_idx + 1, value=min_price_col)
        print(f"  新增列: {min_price_col} (第{min_price_col_idx + 1}列)")
    col_indices[min_price_col] = min_price_col_idx
    
    # 设置买入后最低价列宽
    min_price_col_letter = openpyxl.utils.get_column_letter(min_price_col_idx + 1)
    ws.column_dimensions[min_price_col_letter].width = 18
    
    # 处理"买入后最低价时间"列（日期列）
    min_price_time_col = "买入后最低价时间"
    if min_price_time_col in headers:
        min_price_time_col_idx = headers.index(min_price_time_col)
    else:
        min_price_time_col_idx = len(headers)
        headers.append(min_price_time_col)
        ws.cell(row=1, column=min_price_time_col_idx + 1, value=min_price_time_col)
        print(f"  新增列: {min_price_time_col} (第{min_price_time_col_idx + 1}列)")
    col_indices[min_price_time_col] = min_price_time_col_idx
    
    # 设置买入后最低价时间列宽
    min_price_time_col_letter = openpyxl.utils.get_column_letter(min_price_time_col_idx + 1)
    ws.column_dimensions[min_price_time_col_letter].width = 18
    
    # 处理"卖出价"列（数字列）
    sell_price_col = "卖出价"
    if sell_price_col in headers:
        sell_price_col_idx = headers.index(sell_price_col)
    else:
        sell_price_col_idx = len(headers)
        headers.append(sell_price_col)
        ws.cell(row=1, column=sell_price_col_idx + 1, value=sell_price_col)
        print(f"  新增列: {sell_price_col} (第{sell_price_col_idx + 1}列)")
    col_indices[sell_price_col] = sell_price_col_idx
    
    # 设置卖出价列宽
    sell_price_col_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
    ws.column_dimensions[sell_price_col_letter].width = 15
    
    # 处理"卖出时间"列（日期列）
    sell_time_col = "卖出时间"
    if sell_time_col in headers:
        sell_time_col_idx = headers.index(sell_time_col)
    else:
        sell_time_col_idx = len(headers)
        headers.append(sell_time_col)
        ws.cell(row=1, column=sell_time_col_idx + 1, value=sell_time_col)
        print(f"  新增列: {sell_time_col} (第{sell_time_col_idx + 1}列)")
    col_indices[sell_time_col] = sell_time_col_idx
    
    # 设置卖出时间列宽
    sell_time_col_letter = openpyxl.utils.get_column_letter(sell_time_col_idx + 1)
    ws.column_dimensions[sell_time_col_letter].width = 15
    
    # 处理"卖出后盈亏(%)"列（百分比列）
    sell_profit_col = "卖出后盈亏(%)"
    if sell_profit_col in headers:
        sell_profit_col_idx = headers.index(sell_profit_col)
    else:
        sell_profit_col_idx = len(headers)
        headers.append(sell_profit_col)
        ws.cell(row=1, column=sell_profit_col_idx + 1, value=sell_profit_col)
        print(f"  新增列: {sell_profit_col} (第{sell_profit_col_idx + 1}列)")
    col_indices[sell_profit_col] = sell_profit_col_idx
    
    # 设置卖出后盈亏列宽
    sell_profit_col_letter = openpyxl.utils.get_column_letter(sell_profit_col_idx + 1)
    ws.column_dimensions[sell_profit_col_letter].width = 15
    
    # 处理"卖出盈亏金额"列（金额列）
    sell_profit_amount_col = "卖出盈亏金额"
    if sell_profit_amount_col in headers:
        sell_profit_amount_col_idx = headers.index(sell_profit_amount_col)
    else:
        sell_profit_amount_col_idx = len(headers)
        headers.append(sell_profit_amount_col)
        ws.cell(row=1, column=sell_profit_amount_col_idx + 1, value=sell_profit_amount_col)
        print(f"  新增列: {sell_profit_amount_col} (第{sell_profit_amount_col_idx + 1}列)")
    col_indices[sell_profit_amount_col] = sell_profit_amount_col_idx
    
    # 设置卖出盈亏金额列宽
    sell_profit_amount_col_letter = openpyxl.utils.get_column_letter(sell_profit_amount_col_idx + 1)
    ws.column_dimensions[sell_profit_amount_col_letter].width = 15
    
    # 处理"卖出后最高价(%)"列（百分比列）
    sell_max_price_col = "卖出后最高价(%)"
    if sell_max_price_col in headers:
        sell_max_price_col_idx = headers.index(sell_max_price_col)
    else:
        sell_max_price_col_idx = len(headers)
        headers.append(sell_max_price_col)
        ws.cell(row=1, column=sell_max_price_col_idx + 1, value=sell_max_price_col)
        print(f"  新增列: {sell_max_price_col} (第{sell_max_price_col_idx + 1}列)")
    col_indices[sell_max_price_col] = sell_max_price_col_idx
    
    # 设置卖出后最高价列宽
    sell_max_price_col_letter = openpyxl.utils.get_column_letter(sell_max_price_col_idx + 1)
    ws.column_dimensions[sell_max_price_col_letter].width = 18
    
    # 处理"卖出后最低价(%)"列（百分比列）
    sell_min_price_col = "卖出后最低价(%)"
    if sell_min_price_col in headers:
        sell_min_price_col_idx = headers.index(sell_min_price_col)
    else:
        sell_min_price_col_idx = len(headers)
        headers.append(sell_min_price_col)
        ws.cell(row=1, column=sell_min_price_col_idx + 1, value=sell_min_price_col)
        print(f"  新增列: {sell_min_price_col} (第{sell_min_price_col_idx + 1}列)")
    col_indices[sell_min_price_col] = sell_min_price_col_idx
    
    # 设置卖出后最低价列宽
    sell_min_price_col_letter = openpyxl.utils.get_column_letter(sell_min_price_col_idx + 1)
    ws.column_dimensions[sell_min_price_col_letter].width = 18
    
    # 处理"卖出后最低价时间"列（日期列）
    sell_min_price_time_col = "卖出后最低价时间"
    if sell_min_price_time_col in headers:
        sell_min_price_time_col_idx = headers.index(sell_min_price_time_col)
    else:
        sell_min_price_time_col_idx = len(headers)
        headers.append(sell_min_price_time_col)
        ws.cell(row=1, column=sell_min_price_time_col_idx + 1, value=sell_min_price_time_col)
        print(f"  新增列: {sell_min_price_time_col} (第{sell_min_price_time_col_idx + 1}列)")
    col_indices[sell_min_price_time_col] = sell_min_price_time_col_idx
    
    # 设置卖出后最低价时间列宽
    sell_min_price_time_col_letter = openpyxl.utils.get_column_letter(sell_min_price_time_col_idx + 1)
    ws.column_dimensions[sell_min_price_time_col_letter].width = 18
    
    # 处理"卖出后涨跌(%)"列（百分比列）
    sell_profit_pct_col = "卖出后涨跌(%)"
    if sell_profit_pct_col in headers:
        sell_profit_pct_col_idx = headers.index(sell_profit_pct_col)
    else:
        sell_profit_pct_col_idx = len(headers)
        headers.append(sell_profit_pct_col)
        ws.cell(row=1, column=sell_profit_pct_col_idx + 1, value=sell_profit_pct_col)
        print(f"  新增列: {sell_profit_pct_col} (第{sell_profit_pct_col_idx + 1}列)")
    col_indices[sell_profit_pct_col] = sell_profit_pct_col_idx
    
    # 设置卖出后涨跌列宽
    sell_profit_pct_col_letter = openpyxl.utils.get_column_letter(sell_profit_pct_col_idx + 1)
    ws.column_dimensions[sell_profit_pct_col_letter].width = 15
    
    # 处理"卖出后涨跌时间"列（日期列）
    sell_profit_pct_time_col = "卖出后涨跌时间"
    if sell_profit_pct_time_col in headers:
        sell_profit_pct_time_col_idx = headers.index(sell_profit_pct_time_col)
    else:
        sell_profit_pct_time_col_idx = len(headers)
        headers.append(sell_profit_pct_time_col)
        ws.cell(row=1, column=sell_profit_pct_time_col_idx + 1, value=sell_profit_pct_time_col)
        print(f"  新增列: {sell_profit_pct_time_col} (第{sell_profit_pct_time_col_idx + 1}列)")
    col_indices[sell_profit_pct_time_col] = sell_profit_pct_time_col_idx
    
    # 设置卖出后涨跌时间列宽
    sell_profit_pct_time_col_letter = openpyxl.utils.get_column_letter(sell_profit_pct_time_col_idx + 1)
    ws.column_dimensions[sell_profit_pct_time_col_letter].width = 18
    
    # 更新数据
    updated_count = 0
    for row_data in result_rows:
        # 找到匹配的代码
        stock_code = None
        for key in row_data:
            if '代' in key and '码' in key:
                stock_code = str(row_data[key]).strip()
                break
        
        if not stock_code or stock_code not in code_to_row:
            continue
        
        row_idx = code_to_row[stock_code]
        
        # 更新各列
        for col_name in new_cols:
            # 从结果中查找对应字段
            value = None
            for key in row_data:
                if col_name in key or key in col_name:
                    value = row_data[key]
                    break
            
            if value:
                col_idx = col_indices[col_name]
                ws.cell(row=row_idx, column=col_idx + 1, value=value)
        
        # 设置下拉列默认值（如果为空）
        for col_name, options, default in dropdown_cols:
            col_idx = col_indices[col_name]
            current_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if not current_value:
                ws.cell(row=row_idx, column=col_idx + 1, value=default)
        
        updated_count += 1
    
    # 为所有数据行设置下拉列默认值（如果还没有）
    for row_idx in range(2, ws.max_row + 1):
        for col_name, options, default in dropdown_cols:
            col_idx = col_indices[col_name]
            current_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if not current_value:
                ws.cell(row=row_idx, column=col_idx + 1, value=default)
    
    # 计算"买入后涨跌(%)"列的公式
    # 找到"现价"列
    current_price_col_idx = None
    for idx, h in enumerate(headers):
        if h and '现' in str(h) and '价' in str(h):
            current_price_col_idx = idx
            break
    
    if current_price_col_idx is not None and buy_price_col_idx is not None:
        current_price_letter = openpyxl.utils.get_column_letter(current_price_col_idx + 1)
        buy_price_letter = openpyxl.utils.get_column_letter(buy_price_col_idx + 1)
        profit_letter = openpyxl.utils.get_column_letter(profit_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            # 只有当"当日买入价格"有值时才计算公式
            buy_price = ws.cell(row=row_idx, column=buy_price_col_idx + 1).value
            if buy_price:
                # 公式: (现价 - 买入价) / 买入价 * 100
                formula = f'=({current_price_letter}{row_idx}-{buy_price_letter}{row_idx})/{buy_price_letter}{row_idx}*100'
                ws.cell(row=row_idx, column=profit_col_idx + 1, value=formula)
    
    # 计算"买入时间"列的公式
    # 当"当日买入价格"有值时显示今天日期，否则为空
    if buy_price_col_idx is not None and buy_time_col_idx is not None:
        buy_price_letter = openpyxl.utils.get_column_letter(buy_price_col_idx + 1)
        buy_time_letter = openpyxl.utils.get_column_letter(buy_time_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            buy_price = ws.cell(row=row_idx, column=buy_price_col_idx + 1).value
            if buy_price:
                # 公式: IF(买入价<>"", TODAY(), "")
                formula = f'=IF({buy_price_letter}{row_idx}<>"",TODAY(),"")'
                ws.cell(row=row_idx, column=buy_time_col_idx + 1, value=formula)
    
    # 计算"卖出后盈亏(%)"列的公式
    # 公式: (卖出价 - 买入价) / 买入价 * 100
    if sell_price_col_idx is not None and buy_price_col_idx is not None and sell_profit_col_idx is not None:
        sell_price_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
        buy_price_letter = openpyxl.utils.get_column_letter(buy_price_col_idx + 1)
        sell_profit_letter = openpyxl.utils.get_column_letter(sell_profit_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            sell_price = ws.cell(row=row_idx, column=sell_price_col_idx + 1).value
            buy_price = ws.cell(row=row_idx, column=buy_price_col_idx + 1).value
            if sell_price and buy_price:
                # 公式: (卖出价 - 买入价) / 买入价 * 100
                formula = f'=({sell_price_letter}{row_idx}-{buy_price_letter}{row_idx})/{buy_price_letter}{row_idx}*100'
                ws.cell(row=row_idx, column=sell_profit_col_idx + 1, value=formula)
    
    # 计算"卖出盈亏金额"列的公式
    # 公式: 卖出价 - 买入价
    if sell_price_col_idx is not None and buy_price_col_idx is not None and sell_profit_amount_col_idx is not None:
        sell_price_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
        buy_price_letter = openpyxl.utils.get_column_letter(buy_price_col_idx + 1)
        sell_profit_amount_letter = openpyxl.utils.get_column_letter(sell_profit_amount_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            sell_price = ws.cell(row=row_idx, column=sell_price_col_idx + 1).value
            buy_price = ws.cell(row=row_idx, column=buy_price_col_idx + 1).value
            if sell_price and buy_price:
                # 公式: 卖出价 - 买入价
                formula = f'={sell_price_letter}{row_idx}-{buy_price_letter}{row_idx}'
                ws.cell(row=row_idx, column=sell_profit_amount_col_idx + 1, value=formula)
    
    # 计算"卖出时间"列的公式
    # 当"卖出价"有值时显示今天日期，否则为空
    if sell_price_col_idx is not None and sell_time_col_idx is not None:
        sell_price_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
        sell_time_letter = openpyxl.utils.get_column_letter(sell_time_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            sell_price = ws.cell(row=row_idx, column=sell_price_col_idx + 1).value
            if sell_price:
                # 公式: IF(卖出价<>"", TODAY(), "")
                formula = f'=IF({sell_price_letter}{row_idx}<>"",TODAY(),"")'
                ws.cell(row=row_idx, column=sell_time_col_idx + 1, value=formula)
    
    # 计算"卖出后涨跌(%)"列的公式
    # 公式: (现价 - 卖出价) / 卖出价 * 100
    if current_price_col_idx is not None and sell_price_col_idx is not None and sell_profit_pct_col_idx is not None:
        current_price_letter = openpyxl.utils.get_column_letter(current_price_col_idx + 1)
        sell_price_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
        sell_profit_pct_letter = openpyxl.utils.get_column_letter(sell_profit_pct_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            sell_price = ws.cell(row=row_idx, column=sell_price_col_idx + 1).value
            if sell_price:
                # 公式: (现价 - 卖出价) / 卖出价 * 100
                formula = f'=({current_price_letter}{row_idx}-{sell_price_letter}{row_idx})/{sell_price_letter}{row_idx}*100'
                ws.cell(row=row_idx, column=sell_profit_pct_col_idx + 1, value=formula)
    
    # 计算"卖出后涨跌时间"列的公式
    # 当"卖出价"有值时显示今天日期，否则为空
    if sell_price_col_idx is not None and sell_profit_pct_time_col_idx is not None:
        sell_price_letter = openpyxl.utils.get_column_letter(sell_price_col_idx + 1)
        
        for row_idx in range(2, ws.max_row + 1):
            sell_price = ws.cell(row=row_idx, column=sell_price_col_idx + 1).value
            if sell_price:
                # 公式: IF(卖出价<>"", TODAY(), "")
                formula = f'=IF({sell_price_letter}{row_idx}<>"",TODAY(),"")'
                ws.cell(row=row_idx, column=sell_profit_pct_time_col_idx + 1, value=formula)
    
    wb.save(source_file)
    print(f"\n已更新 {updated_count} 条记录到源文件 {source_file}")


def main():
    # 1. 读取个股信息（代码和名称）
    stock_info = read_stock_info(INPUT_FILE)
    print(f"共读取 {len(stock_info)} 只个股")

    all_result_rows = []

    # 2. 分批查询
    total_batches = (len(stock_info) + BATCH_SIZE - 1) // BATCH_SIZE
    for i in range(0, len(stock_info), BATCH_SIZE):
        batch = stock_info[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        batch_names = [name for code, name in batch]
        print(f"\n[{batch_num}/{total_batches}] 查询第 {i+1}-{i+len(batch)} 只个股: {', '.join(batch_names)}")

        try:
            result_text = query_stocks(batch)
            print(f"  解析结果...")
            rows = parse_response(result_text)
            print(f"  解析到 {len(rows)} 条记录")
            if rows:
                all_result_rows.extend(rows)
                # 每批查询后实时更新到源文件
                update_source_file(all_result_rows, INPUT_FILE)
        except Exception as e:
            print(f"  ❌ 查询失败: {type(e).__name__}: {e}")

        # 请求间隔，避免频率过高
        if i + BATCH_SIZE < len(stock_info):
            time.sleep(random.randint(15, 30))

    # 3. 最终更新
    update_source_file(all_result_rows, INPUT_FILE)
    print("\n全部完成！")


if __name__ == "__main__":
    main()
